#!/usr/local/bin/python -u

"""
NFL stats: a program to calculate NFL team standings

Script is now designed as a module. To get started, use the command line options or load
interactively and try:

from nfl import NFL
nfl = NFL()
nfl.load('NFLData.xlsx')

nfl('MIN')                       # info and standings for vikings
nfl('NFC')                       # or the NFC conference
nfl.wlt(['MIN', 'GB'])           # win/loss/tie info
nfl.tiebreakers(['DAL', 'PHI'])  # tiebreaker calculations

# and lots more, and yet lots of room for enhancements/improvements ;)

Usage:
    nfl.py update FILE WEEK
    nfl.py team TEAM [--file FILE]
    nfl.py tiebreakers TEAMS... [--file FILE]

Commands:
    update          Update the database FILE, scraping data up to and including WEEK

    team            Report TEAM stats and schedule (division or conference work too)

    tiebreakers     Report tiebreaker analysis for the specified TEAMS

Options:
    --file FILE     Use FILE as database path [default: NFLData.xlsx]

"""


import openpyxl
import pprint
import requests
import sys
import re
import logging
import time
from docopt import docopt
from pyquery import PyQuery
import urllib
from datetime import datetime
import numpy as np
import pandas as pd

class NFLTeam():
    def __init__(self, code, host):

        elem = host.teams_[code]
        self.code = code
        self.name = elem['name']
        self.div  = elem['div']
        self.conf = elem['conf']
        self.host = host

    @property
    def division(self):
        '''Return array of teams in this team's division
        '''
        return self.host.divs_[self.div]

    @property
    def conference(self):
        '''Return array of teams in this team's conference
        '''
        return self.host.confs_[self.conf]

    @property
    def schedule(self):
        '''Return the team's schedule
        '''

        return self.host.schedule(self.code)

    def __repr__(self):
        z = self.host._stats().loc[self.code][['overall','division','conference']].unstack()
        return '{}: {} ({})\n'.format(self.code, self.name, self.div) + z.__repr__()


class NFLDivision():
    def __init__(self, code, host):

        self.code = code
        self.host = host
        self.teams = host.divs_[code]

    def standings(self, rank=False):
        ''' Return division standings

            rank: if True, add 'rank' column based on in-division tiebreaker rules (longer)
                  if False, sort results by game wlt
        '''

        o = self.host.wlt(self.teams)
        d = self.host.wlt(self.teams, within=self.teams)
        z = pd.concat([o, d], keys=['overall', 'division'], axis=1)

        # use division tiebreakers to calculate precise division rank
        if rank:
            t = self.host.tiebreakers(self.teams, ascending=True).xs('pct', axis=1, level=1).transpose()
            t = t.sort_values(list(t.columns), ascending=False)
            t['div_rank'] = range(1, len(t)+1)
            z[('division','rank')] = t['div_rank']
            return z.sort_values(('division','rank'))

        return z

    def __repr__(self):
        z = self.host._stats()
        z = z[z['div']==self.code][['overall','division']]
        return '{}\n'.format(self.code) + z.__repr__()

class NFLConference():
    def __init__(self, code, host):

        self.code = code
        self.host = host
        self.teams = host.confs_[code]

    @property
    def divisions(self):
        return set(map(lambda x: '-'.join([self.code, x]), ['North', 'East', 'South', 'West']))


    def standings(self, rank=False):

        c = None
        for elem in self.divisions:
            z = self.host(elem).standings(rank)
            if type(c) is pd.DataFrame:
                c = pd.concat([c, z])
            else:
                c = z

        z = self.host.wlt(self.teams, within=self.teams)
        c = pd.concat([c['overall'], c['division'], z], keys=['overall', 'division', 'conference'], axis=1)
        c['div'] = c.index.map(lambda x: self.host.teams_[x]['div'].split('-')[1])
        if rank:
            return c.sort_values(['div', ('division','rank')])

        return c.sort_values(['div',('overall','pct')], ascending=[True, False])

    def __repr__(self):
        z = self.host._stats()
        return '{}\n'.format(self.code) + z[z['conf']==self.code][['div', 'overall', 'division', 'conference']].__repr__()

class NFL():

    def __init__(self):
        self.teams_  = {}
        self.divs_   = {}
        self.confs_  = {}
        self.games_  = []
        self.path   = None
        self.max_week = 0
        self.stats = pd.DataFrame()

    def __call__(self, i):

        if i in self.confs_:
            return NFLConference(i, self)

        if i in self.divs_:
            return NFLDivision(i, self)

        if i in self.teams_:
            return NFLTeam(i, self)

    def load(self, path='NFLData.xlsx'):
        ''' Loads data from the specified excel file

            path:   path to Excel file
        '''

        # in case of reload
        self.path   = path
        self.games_ = []
        self.teams_ = {}
        self.divs_  = {}
        self.confs_ = {}
        self.max_week = 0                  

        wb = openpyxl.load_workbook(path, read_only=True)
        for row in wb['Divisions']:
            team = row[0].value
            conf = row[2].value
            div = row[3].value
            if team and conf and div:
                div = '-'.join([conf, div])
                self.teams_[team] = {'name': row[1].value, 'div': div, 'conf': conf}
                if self.divs_.get(div) is None:
                    self.divs_[div] = set()

                self.divs_[div].add(team)

                if self.confs_.get(conf) is None:
                    self.confs_[conf] = set()

                self.confs_[conf].add(team)


        for row in wb['Scores']:
            if row[0].row > 1 and row[0].value:
                # at/ht = away team/home team - same for scores
                game = {'wk': row[0].value, 'at': row[1].value, 'as': row[2].value, 'ht': row[3].value, 'hs': row[4].value}
                game['p'] = game['as'] is not None and game['hs'] is not None
                self.games_.append(game)
                self.max_week = max(self.max_week, game['wk'])

        self.stats = None
        return self

    def _stats(self):

        if self.stats is not None:
            return self.stats

        stat_cols = [('name',''),('div',''),('conf','')]
        stat_cols += list(pd.MultiIndex.from_product([['overall','division','conference', 'vic_stren', 'sch_stren'],['win','loss','tie','pct']]))
        stat_cols += list(pd.MultiIndex.from_product([['misc'],['rank-conf','rank-overall', 'pts-scored', 'pts-allowed']]))
        stats = pd.DataFrame(columns=pd.MultiIndex.from_tuples(stat_cols))

        sched = {k:[] for k in self.teams_.keys()}

        def tally(hcode, acode, hscore, ascore, cat):

            if hscore > ascore:
                stats.loc[hcode, (cat,'win')] += 1
                stats.loc[acode, (cat,'loss')] += 1
                if cat == 'overall':
                    sched[hcode] += [(acode,1)]
                    sched[acode] += [(hcode,0)]
            elif hscore < ascore:
                stats.loc[hcode, (cat,'loss')] += 1
                stats.loc[acode, (cat,'win')] += 1
                if cat == 'overall':
                    sched[hcode] += [(acode,0)]
                    sched[acode] += [(hcode,1)]
            elif hscore == ascore:
                stats.loc[hcode, (cat,'tie')] += 1
                stats.loc[acode, (cat,'tie')] += 1
                if cat == 'overall':
                    sched[hcode] += [(acode,0)]
                    sched[acode] += [(hcode,0)]                   

        for k,team in self.teams_.items():
            stats.loc[k, ['name', 'div', 'conf']] = (team['name'], team['div'], team['conf'])
            stats.loc[k, ['overall','division','conference', 'vic_stren', 'sch_stren', 'misc']] = 0

        for game in self.games_:
            if game['p']:
                tally(game['ht'], game['at'], game['hs'], game['as'], 'overall')
                stats.loc[game['ht'], ('misc', 'pts-scored')] += game['hs']
                stats.loc[game['at'], ('misc', 'pts-scored')] += game['as']
                stats.loc[game['ht'], ('misc', 'pts-allowed')] += game['as']
                stats.loc[game['at'], ('misc', 'pts-allowed')] += game['hs']

                if self.teams_[game['ht']]['div'] == self.teams_[game['at']]['div']:
                    tally(game['ht'], game['at'], game['hs'], game['as'], 'division')
                    tally(game['ht'], game['at'], game['hs'], game['as'], 'conference')
                elif self.teams_[game['ht']]['conf'] == self.teams_[game['at']]['conf']:
                    tally(game['ht'], game['at'], game['hs'], game['as'], 'conference')

        # strength of victory/schedule
        for (k,row) in stats.iterrows():
            for (op,win) in sched[k]:
                stats.loc[k, 'sch_stren'] = (stats.loc[k, 'sch_stren'] + stats.loc[op, 'overall']).values
                if win == 1:
                    stats.loc[k, 'vic_stren'] = (stats.loc[k, 'vic_stren'] + stats.loc[op, 'overall']).values

        # temporary table for calculating ranks - easier syntax
        t = pd.concat([stats['conf'], stats['misc'][['pts-scored','pts-allowed']]], axis=1)
        stats[('misc','rank-overall')] = (t['pts-scored'].rank() + t['pts-allowed'].rank(ascending=False)).rank()

        t['conf-off-rank'] = t.groupby('conf')['pts-scored'].rank()
        t['conf-def-rank'] = t.groupby('conf')['pts-allowed'].rank(ascending=False)
        t['conf-rank'] = t['conf-off-rank'] + t['conf-def-rank']
        stats[('misc', 'rank-conf')] = t.groupby('conf')['conf-rank'].rank()

        for i in ['overall','division','conference', 'sch_stren', 'vic_stren']:
            stats[(i,'pct')] = (stats[(i,'win')] + stats[(i,'tie')]*0.5) / stats[i].sum(axis=1)

        stats.sort_values(['div',('overall','pct')], ascending=(True,False), inplace=True)
        self.stats = stats
        return self.stats


    def reload(self):
        ''' Reloads the previous Excel file
        '''

        if self.path:
            self.load(self.path)

        return self

    def update(self, path, week, year=None):
        ''' Updates the Scores sheet of the specified Excel file by scraping the URL shown in code.
            The file must already exist and have a valid "Divisions" sheet with team names consistent
            with those on the source website.

            path:   path to Excel workbook

            week:   week number (1-based) of last week to load

            year:   season to load; else infer the latest season from the current date
        '''

        if not year:
            dt = datetime.now()
            year = dt.year if dt.month >= 8 else dt.year-1

        self.load(path)
        tids = {v['name']:k for k,v in self.teams_.items()}
        wb = openpyxl.load_workbook(path, read_only=False)
        ws = wb['Scores']
        row = 2

        def safeInt(i):
            try:
                i = int(i)
            except ValueError:
                pass

            return i

        for w in range(1, week+1):
            url = 'https://www.pro-football-reference.com/years/{}/week_{}.htm'.format(year, w)
            try:
                time.sleep(1) # avoid throttling
                d = PyQuery(url=url)('div.game_summaries div.game_summary table.teams')
            except urllib.error.HTTPError as err:
                logging.error('Bad URL: {}'.format(url))
                raise

            logging.info('Processing {}/{}: ({}) - {} games'.format(w, year, url, len(d)))
            for elem in d:
                ateam = PyQuery(PyQuery(elem)('tr:nth-child(2)'))
                bteam = PyQuery(PyQuery(elem)('tr:nth-child(3)'))
                (aname,ascore) = (ateam('td:nth-child(1)').text(), ateam('td:nth-child(2)').text())
                (bname,bscore) = (bteam('td:nth-child(1)').text(), bteam('td:nth-child(2)').text())

                try:
                    aname = tids[aname]
                    bname = tids[bname]
                except KeyError as k:
                    logging.error('{}: unrecognized team name: check the Divisions table'.format(k))
                    raise

                ws.cell(row=row, column=1, value=w)
                ws.cell(row=row, column=2, value=aname)
                ws.cell(row=row, column=3, value=safeInt(ascore))
                ws.cell(row=row, column=4, value=bname)
                ws.cell(row=row, column=5, value=safeInt(bscore))
                row += 1

        # empty the remaining cells
        for row in range(row,ws.max_row):
            for c in range(1,6):
                ws.cell(row=row, column=c, value=None)

        wb.save(path)
        return self

    def set(self, wk, **kwargs):
        ''' Set the final score(s) for games in a given week. You can use this to create
            hypothetical outcomes and analyze the effect on team rankings. Scores
            are specified by team code and applied to the specified week's schedule.
            If the score is specified for only one team in a game, the score of the other
            team is assumed to be zero if not previously set.

            wk:         week number
            **kwargs    dict of team codes and final scores

            # typical example
            set(7, MIN=17, GB=10)

            The above will set the score for the MIN/GB game in week 7 if
            MIN and GB play each other in that week. Otherwise, it will
            set each team's score respectively and default their opponent's scores
            to 0. Scores for bye teams in that week are ignored.
        '''

        # sanity checks
        bogus = set(kwargs.keys()) - set(self.teams_.keys())
        if len(bogus) > 0:
            raise KeyError('Invalid team codes: {}'.format(','.join(bogus)))       

        for elem in self.games_:
            if wk == elem['wk']:
                t = False
                if elem['ht'] in kwargs:
                    elem['hs'] = kwargs[elem['ht']]
                    t = True

                if elem['at'] in kwargs:
                    elem['as'] = kwargs[elem['at']]
                    t = True

                # sanity checks
                if t:
                    elem['hs'] = elem['hs'] or 0
                    elem['as'] = elem['as'] or 0
                    elem['p'] = True

            elif wk < elem['wk']:
                # assuming elements are sorted by week, we can stop at this point
                break

        self.stats = None # signal to rebuild stats
        return self


    def games(self, teams=None, limit=None, allGames=False):
        ''' generator to iterate over score data

            teams:      code or list-like of teams to fetch

            limit:      range or list-like of weeks to fetch. Integers are converted
                        to the top limit of a range

            Example:

                for score in scores('MIN', limit=10) # fetch Vikings record up to but not including week 10
        '''

        if type(limit) is int:
            limit = range(1, limit)

        teams = self._teams(teams)

        for elem in self.games_:
            if limit is None or elem['wk'] in limit:
                if teams is None or elem['at'] in teams or elem['ht'] in teams:
                    if elem['p'] or allGames:
                        yield elem


    def scores(self, teams=None, limit=None):
        ''' Returns interated game data structured by teams

            Result is  dict keyed by team code each of game results as follows:

            [wlt, us, them, op, home, week]

            wlt:  'win' 'loss' or 'tie'
            us:   our final score
            them: their final score
            op:   opponent (team code)
            home: True for home games
            week: week number
        '''

        if teams is None:
            z = {i:[] for i in self.teams_.keys()}
        else:
            teams = self._teams(teams)
            z = {i:[] for i in teams}

        for game in self.games(teams, limit):
            if teams is None or game['at'] in teams:
                z[game['at']].append([NFL.result(game['as'], game['hs']), game['as'], game['hs'], game['ht'], False, game['wk']])

            if teams is None or game['ht'] in teams:
                z[game['ht']].append([NFL.result(game['hs'], game['as']), game['hs'], game['as'], game['at'], True, game['wk']])

        return z


    def schedule(self, which):

        if type(which) is int:
            df = pd.DataFrame(columns=['ht', 'at', 'hscore', 'ascore'])
            for game in self.games(teams=None, limit=range(which,which+1), allGames=True):
                df.loc[len(df)] = [game['ht'], game['at'], game['hs'], game['as']]
        else:
            # pre-populate the index so the schedule includes the bye week
            df = pd.DataFrame(columns=['opp', 'at_home', 'score', 'opp_score', 'wlt'], index=range(1, self.max_week+1))
            df.index.name = 'week'
            for game in self.games(teams=which, allGames=True):
                if game['ht'] == which:
                    df.loc[game['wk']] = [game['at'], 1, game['hs'], game['as'], NFL.result(game['hs'],game['as'])]
                else:
                    df.loc[game['wk']] = [game['ht'], 0, game['as'], game['hs'], NFL.result(game['as'], game['hs'])]

        return df

    def opponents(self, teams, limit=None):
        ''' Returns the set of common opponents of one or more teams

            The teams argument can be a single team or a list.
        '''

        if teams is None:
            raise ValueError("teams cannot be None here")

        teams = self._teams(teams)

        ops = {t:set() for t in teams}

        for game in self.games(teams, limit):
            if game['ht'] in ops:
                ops[game['ht']].add(game['at'])

            if game['at'] in ops:
                ops[game['at']].add(game['ht'])

        # Resulting set is common (intersection) of opponents excluding the teams themselves
        z = None
        for s in ops.values():
            if z is None:
                z = s
            else:
                z &= s

        z -= set(teams)
        return z


    def wlt(self, teams=None, within=None, limit=None):
        ''' Return the wlt stats of one or more teams

            teams:  team code or list of team codes

            within: list of team codes that defines the wlt universe
        '''

        teams = self._teams(teams)

        df = pd.DataFrame(columns=['win','loss','tie'])
        df.columns.name = 'outcome'
        df.index.name = 'team'
        for t,scores in self.scores(teams, limit).items():
            df.loc[t] = [0, 0, 0]
            for score in scores:
                if within is None or score[3] in within:
                    df.loc[t, score[0]] += 1

        df['pct'] = (df['win'] + df['tie'] * 0.5) / df.T.sum()
        return df.sort_values('pct', ascending=False)

    def team_stats(self, team, limit=None):

        df = pd.concat([
            self.wlt(team, limit=limit).rename(index={team: 'overall'}),
            self.wlt(team, within=self(team).division, limit=limit).rename(index={team: 'division'}),
            self.wlt(team, within=self(team).conference, limit=limit).rename(index={team: 'conference'})
        ])

        df.index.name = None
        return df


    def point_rankings(self, limit=None, ascending=False):
        ''' Return overall and conference rankings of points scored and points allowed. These
            are used in tiebreakers

            ascending: if True, best-performing teams have the highest number rankings
        '''
        df = pd.DataFrame()
        for k,elem in self.teams_.items():
            df.loc[k, 'conf'] = elem['conf']

        df['pts_scored'] = 0
        df['pts_allowed'] = 0

        for elem in self.games(limit):
            df.loc[elem['ht'], ['pts_scored','pts_allowed']] += [elem['hs'], elem['as']]
            df.loc[elem['at'], ['pts_scored','pts_allowed']] += [elem['as'], elem['hs']]

        df['rank_scored'] = df['pts_scored'].rank(ascending=ascending)
        df['rank_allowed'] = df['pts_allowed'].rank(ascending=not ascending)
        df['rank_overall'] = (df['rank_scored'] + df['rank_allowed']).rank(ascending=True)
        df['rank_conf'] = df.groupby('conf')['rank_overall'].rank(ascending=True)

        return df


    def tiebreakers(self, teams, limit=None, ascending=False):
        '''Return tiebreaker analysis for specified teams

        Each row in the returned dataframe is the results of a step in the NFL's tiebreaker procedure
        currently defined here: https://www.nfl.com/standings/tie-breaking-procedures

        Rows are in order of precedence and depend on whether the teams are in the same division or not

        If ascending is True, "best combined rankings" mean that higher values are better, which enables
        sorting like this:

        z = nfl.tiebreakers(nfl('NFC-North').teams, ascending=True)
        z = z.xs('pct', level=1, axis=1).transpose()
        z.sort_values(list(z.columns))     # lowest to highest
        '''

        teams = self._teams(teams)

        df = pd.DataFrame(columns=pd.MultiIndex.from_product([teams, ['win','loss','tie', 'pct']], names=['team','outcome']))
        divisions = set()

        # determine which divisions are in the specified list. If more than one then adjust the tiebreaker order
        for t in teams:
            divisions.add(self.teams_[t]['div'])

        def set_frame(df, index, z):
            '''Assigns the index and columns of a data frame to a multi-indexed data frame with the same column levels.
               There might be a simpler way to do this, but this works okay

               df:      the target dataframe

               index:   index for assignment

               z:       source dataframe. z.index should equal df.columns[0] and z.columns should equal df.columns[1]
            '''

            df.loc[index] = pd.Series(z.values.flatten(), index=pd.MultiIndex.from_product([z.index, z.columns]))

        rankings = self.point_rankings(limit=limit, ascending=ascending)
        common_opponents = self.opponents(teams, limit=limit)
        scores = self.scores(teams, limit=limit)

        # declare these rows in advance so they appear in the correct order
        set_frame(df, 'overall', self.wlt(teams, limit=limit))
        set_frame(df, 'head-to-head', self.wlt(teams, within=teams, limit=limit))

        if len(divisions) > 1:
            df.loc['conference'] = np.nan
            df.loc['common-games'] = np.nan
        else:
            df.loc['division'] = np.nan
            df.loc['common-games'] = np.nan
            df.loc['conference'] = np.nan
        
        df.loc['victory-strength'] = np.nan
        df.loc['schedule-strength'] = np.nan
        df.loc['conference-rank'] = np.nan
        df.loc['overall-rank'] = np.nan
        df.loc['points-common-games'] = np.nan
        df.loc['points-overall'] = np.nan

        set_frame(df, 'common-games', self.wlt(teams, within=common_opponents, limit=limit))

        for t in teams:
            if 'division' in df.index:
                df.loc['division', t] = list(self.wlt(t, within=self(t).division, limit=limit).T[t])

            df.loc['conference', t] = list(self.wlt(t, within=self(t).conference, limit=limit).T[t])

            df.loc['conference-rank', (t, 'pct')] = rankings.loc[t]['rank_conf']
            df.loc['overall-rank', (t, 'pct')] = rankings.loc[t]['rank_overall']

            df.loc['victory-strength', t] = 0
            df.loc['schedule-strength', t] = 0

            sched_wlt = self.wlt(self.opponents(t, limit=limit))

            pts_common_games = pts_overall = 0

            # See https://stackoverflow.com/questions/28431519/pandas-multiindex-assignment-from-another-dataframe
            # for why the syntax below is so convoluted
            for score in scores[t]:
                pts_overall += score[1]
                if score[3] in common_opponents:
                    pts_common_games += score[1]

                df.loc['schedule-strength', t] = (df.loc['schedule-strength', t] + sched_wlt.loc[score[3]]).values
                if score[0] == 'win':
                    df.loc['victory-strength', t] = (df.loc['victory-strength', t] + sched_wlt.loc[score[3]]).values

            df.loc['points-common-games', (t, 'pct')] = pts_common_games
            df.loc['points-overall', (t, 'pct')] = pts_overall


        # recalulcate the pct's we summed above
        # NB: pandas is not entirely consistent concerning when .loc returns a view or a copy
        # See https://stackoverflow.com/questions/26879073/checking-whether-data-frame-is-copy-or-view-in-pandas
        # for more information (and how to tell when it's a view or a copy)
        # The code below assumes .loc returns a copy even though the slice operator returns
        # a view at the time this code was written

        s = df.loc['victory-strength':'schedule-strength'].T

        s.loc(axis=0)[:,'pct'] = 0          # so the formula below doesn't factor in the old pct
        s.loc(axis=0)[:,'pct'] = ((s.xs('win',level=1) + s.xs('tie',level=1)*0.5) / s.groupby(level=0).sum()).values

        # copy back: see note above
        df.loc['victory-strength':'schedule-strength'] = s.T.values

        return df

    @staticmethod
    def result(a, b):
        
        if a is None or b is None:
            return ''
        elif a > b:
            return 'win'
        elif a < b:
            return 'loss'

        return 'tie'

    def _teams(self, teams):
        ''' Transforms teams into an array of actual team codes, or None
        '''

        if teams is None:
            return None

        if type(teams) in [list, set]:
            # already a list
            return teams

        if teams in self.divs_:
            # division code
            return self.divs_[teams]

        if teams in self.confs_:
            # conference code
            return self.confs_[teams]

        # single team code
        return [teams]


if __name__ == '__main__':
    config = docopt(__doc__)
    # print(config)
    # sys.exit(0)

    if config['update']:
        logging.basicConfig(level=logging.INFO)

        NFL().update(config['FILE'], week=int(config['WEEK']))

    if config['team']:
        nfl = NFL()
        nfl.load(config['--file'])
        team = nfl(config['TEAM'])
        print(team)
        if type(team) is NFLTeam:
            print(team.schedule)

    if config['tiebreakers']:
        nfl = NFL()
        print(nfl.load(config['--file']).tiebreakers(config['TEAMS']))


